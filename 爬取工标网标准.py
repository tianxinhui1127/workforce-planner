import requests
from bs4 import BeautifulSoup
import time
import pandas as pd
from fake_useragent import UserAgent
import openpyxl
from openpyxl.styles import Font
import logging
import random

# 配置日志记录
logging.basicConfig(level=logging.INFO, filename='crawler.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 初始化UserAgent
ua = UserAgent()

# 添加更多的User-Agent选项
desktop_agents = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0'
]

# 创建一个空列表来存储数据
data = []

# 定义重试次数和延迟
max_retries = 5

# 使用Session来管理请求，并设置持久性Cookie
session = requests.Session()

# 添加随机延迟函数
def random_delay():
    return random.uniform(15, 30)  # 随机延迟15-30秒

# 循环遍历所有页面，从第1页到第1304页
for page in range(1, 1304):
    # 随机选择User-Agent
    current_ua = random.choice(desktop_agents)
    
    headers = {
        'User-Agent': current_ua,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Referer': f'http://www.csres.com/sort/industry/002006_{max(1, page-1)}.html',  # 添加真实的Referer
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
        'Accept-Encoding': 'gzip, deflate',
        'DNT': '1',  # 添加Do Not Track头
    }
    
    # 在URL中添加随机参数
    timestamp = int(time.time() * 1000)
    url = f'http://www.csres.com/sort/industry/002006_{page}.html?t={timestamp}'

    retry_count = 0
    while retry_count < max_retries:
        try:
            # 添加随机延迟
            delay = random_delay()
            logging.info(f'等待 {delay:.2f} 秒后访问页面 {page}')
            time.sleep(delay)
            
            response = session.get(url, headers=headers, timeout=30)
            if response.status_code == 404:
                logging.warning(f'页面 {page} 不存在，跳过。')
                break  # 跳过该页面
            elif response.status_code != 200:
                logging.error(f'页面 {page} 无法访问，状态码: {response.status_code}')
                break  # 跳过该页面

            soup = BeautifulSoup(response.content, 'lxml')
            rows = soup.find_all('tr', bgcolor="#FFFFFF")
            if not rows:
                logging.warning(f'页面 {page} 中没有数据行，跳过。')
                break

            for row in rows:
                try:
                    tds = row.find_all('td')
                    if len(tds) < 5:
                        logging.warning(f'页面 {page} 中存在不完整的行，跳过。')
                        continue

                    standard_number = tds[0].get_text(strip=True)
                    standard_name = tds[1].get_text(strip=True)
                    standard_status = tds[4].get_text(strip=True)
                    data.append({'标准编号': standard_number, '标准名称': standard_name, '状态': standard_status})
                    logging.info(f'编号: {standard_number}, 名称: {standard_name}, 状态: {standard_status}')

                except Exception as e:
                    logging.error(f'页面 {page} 中提取数据出错: {e}')
                    continue

            # 每抓取10页后增加一个较长的休息时间
            if page % 10 == 0:
                long_delay = random.uniform(60, 120)  # 休息1-2分钟
                logging.info(f'已完成10页抓取，休息 {long_delay:.2f} 秒')
                time.sleep(long_delay)
                
            break  # 请求成功，跳出重试循环

        except requests.exceptions.ReadTimeout:
            logging.warning(f'页面 {page} 请求超时，重试次数 {retry_count + 1}')
            time.sleep(5 * (retry_count + 1))
            retry_count += 1

        except Exception as e:
            logging.error(f'页面 {page} 发生错误: {e}')
            retry_count += 1



# 将数据转换为数据框
df = pd.DataFrame(data)

# 保存到Excel文件
df.to_excel('standards.xlsx', index=False)

# 加载Excel文件并设置字体颜色
try:
    wb = openpyxl.load_workbook('standards.xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    std_name_col = headers.index('标准名称') + 1
    status_col = headers.index('状态') + 1
    for row in ws.iter_rows(min_row=2):
        if len(row) < max(std_name_col, status_col):
            continue
        status_cell = row[status_col - 1]
        if status_cell.value in ['废止', '作废']:
            for cell in row:
                cell.font = Font(color='FF0000')  # 整行红色
    wb.save('standards_with_color.xlsx')
except Exception as e:
    logging.error(f'Excel文件处理出错: {e}')

logging.info('数据爬取完成，结果保存在standards_with_color.xlsx文件中')
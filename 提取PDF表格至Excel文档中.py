import pdfplumber
import pandas as pd
from openpyxl.styles import Border, Side, Alignment, Font
import os
import tkinter as tk
from tkinter import filedialog

def extract_tables_from_pdf(pdf_path):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables()
            if page_tables:
                for table in page_tables:
                    # 提取表格上方的标题
                    title = page.extract_text().split('\n')[0]  # 假设标题在第一页的第一行
                    table.insert(0, [title])  # 将标题插入到表格的第一行
                    tables.append(table)
    return tables

def save_tables_to_excel(tables, excel_path):
    existing_sheets = {}
    for index, table in enumerate(tables):
        # 删除第二行
        if len(table) > 1:  # 确保表格有第二行
            table.pop(1)  # 删除第二行

        df = pd.DataFrame(table)
        # 将工作表名称修改为表格第一行的文本内容
        sheet_name = str(table[0][0])

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if index > 0 else 'w') as writer:
            workbook = writer.book
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                existing_sheets[sheet_name] = worksheet
            else:
                worksheet = workbook.create_sheet(title=sheet_name)
                existing_sheets[sheet_name] = worksheet

            # 设置标题行跨行居中
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 根据第二行的宽度合并第一行
            num_columns = len(table[1]) if len(table) > 1 else 1  # 获取第二行的列数（如果存在）
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

            # 设置标题内容
            title_cell.value = table[0][0]  # 假设标题在表格的第一行第一列

            # 设置字体为宋体12号加粗
            title_cell.font = Font(name='宋体', size=12, bold=True)

            # 设置第一行行高
            worksheet.row_dimensions[1].height = 25.5

            # 设置边框
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # 追加数据
            start_row = worksheet.max_row + 1 if worksheet.max_row > 1 else 2
            for r_idx, row in enumerate(df.values[2:], start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    cell.border = thin_border

            # 调整行高
            for row in worksheet.iter_rows(min_row=start_row):
                max_length = 0
                for cell in row:
                    try:
                        cell_length = len(str(cell.value))
                    except:
                        cell_length = 0
                    max_length = max(max_length, cell_length)
                worksheet.row_dimensions[row[0].row].height = (max_length + 2) * 1.5

        # 确保至少有一个工作表可见
        if index == 0:
            workbook.active = 0

def main():
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 选择PDF文件
    pdf_file = filedialog.askopenfilename(title="选择PDF文件", filetypes=[("PDF files", "*.pdf")])
    if not pdf_file:
        print("未选择PDF文件")
        return

    # 选择保存Excel文件的路径
    excel_file = filedialog.asksaveasfilename(title="保存Excel文件", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not excel_file:
        print("未选择保存路径")
        return

    table_data = extract_tables_from_pdf(pdf_file)
    save_tables_to_excel(table_data, excel_file)
    print(f"表格已成功从 {pdf_file} 转换至 {excel_file}")

if __name__ == "__main__":
    main()